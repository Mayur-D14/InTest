import io
from openpyxl import Workbook, load_workbook

from app import models, schemas

# Required columns, matched case-insensitively with whitespace trimmed.
# "Linked Script" is optional and not part of the required format.
REQUIRED_COLUMNS = [
    "Test Title", "Description", "Priority", "Severity",
    "Test Scripts", "Test Data", "Expected Result", "Actual Result",
]
OPTIONAL_COLUMNS = ["Linked Script"]

PRIORITY_VALUES = {p.value.lower(): p for p in models.Priority}
SEVERITY_VALUES = {s.value.lower(): s for s in models.Severity}


class ExcelParseError(Exception):
    pass


def _normalize_header(h) -> str:
    return str(h).strip() if h is not None else ""


def parse_excel(file_bytes: bytes) -> tuple[list[schemas.TestCaseVersionIn], list[schemas.ExcelUploadRowError]]:
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ExcelParseError(f"Could not read the Excel file: {e}")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ExcelParseError("The uploaded file has no rows.")

    headers = [_normalize_header(h) for h in header_row]
    header_index = {h.lower(): i for i, h in enumerate(headers) if h}

    missing = [c for c in REQUIRED_COLUMNS if c.lower() not in header_index]
    if missing:
        raise ExcelParseError(
            f"Missing required column(s): {', '.join(missing)}. "
            f"Expected format: {' | '.join(REQUIRED_COLUMNS)}"
        )

    def cell(row, col_name: str) -> str:
        idx = header_index.get(col_name.lower())
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    valid_rows: list[schemas.TestCaseVersionIn] = []
    errors: list[schemas.ExcelUploadRowError] = []

    for row_number, row in enumerate(rows_iter, start=2):  # row 1 is the header
        if row is None or all(v is None for v in row):
            continue  # skip fully blank rows

        title = cell(row, "Test Title")
        if not title:
            errors.append(schemas.ExcelUploadRowError(row_number=row_number, message="Test Title is required — row skipped."))
            continue

        priority_raw = cell(row, "Priority").lower()
        priority = PRIORITY_VALUES.get(priority_raw, models.Priority.MEDIUM)

        severity_raw = cell(row, "Severity").lower()
        severity = SEVERITY_VALUES.get(severity_raw, models.Severity.MAJOR)

        linked_script = cell(row, "Linked Script") if "linked script" in header_index else ""

        valid_rows.append(schemas.TestCaseVersionIn(
            title=title,
            description=cell(row, "Description"),
            priority=priority,
            severity=severity,
            test_scripts=cell(row, "Test Scripts"),
            test_data=cell(row, "Test Data"),
            expected_result=cell(row, "Expected Result"),
            actual_result=cell(row, "Actual Result"),
            linked_script_name=linked_script or None,
            change_summary="Imported from Excel",
        ))

    return valid_rows, errors


def build_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    all_columns = REQUIRED_COLUMNS + OPTIONAL_COLUMNS
    ws.append(all_columns)
    ws.append([
        "User can log in with valid credentials",
        "Verifies successful login with a valid username/password combo",
        "High",
        "Major",
        "1. Navigate to /login\n2. Enter valid credentials\n3. Click Sign In",
        "username: testuser@example.com / password: Test1234!",
        "User is redirected to the dashboard",
        "",
        "",
    ])

    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
